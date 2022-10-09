import datetime
from pyexpat import model
from tkinter.ttk import Combobox
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from openpyxl import load_workbook
from tkinter import *
import sys
import os
from sklearn.metrics import r2_score
import ml_metrics as metrics

# file = "Information.xlsx"
# wb = load_workbook(file)
# wb.active = 0
# sheet = wb.active
#
# for i in range(3, len(sheet['B']) + 1):
#     id = str(sheet['A' + str(i)].value)
#     area = str(sheet['B' + str(i)].value)
#     medOrganization = str(sheet['C' + str(i)].value)
#     numKard = str(sheet['D' + str(i)].value)
#     age = str(sheet['E' + str(i)].value)
#     gender = str(sheet['F' + str(i)].value)
#     dataHospital = str(sheet['G' + str(i)].value)
#     datadischarge = str(sheet['H' + str(i)].value)
#     mkb = str(sheet['I' + str(i)].value)
#     diagnosis = str(sheet['J' + str(i)].value)
#     resultHealth = str(sheet['K' + str(i)].value)
#     healtIshod = str(sheet['L' + str(i)].value)
#     typeOfAssistance = str(sheet['L' + str(i)].value)
#     resultKVI = str(sheet['L' + str(i)].value)
#     dataResultKVI = str(sheet['L' + str(i)].value)
# print(area)
file = "Information.xlsx"
wb = load_workbook(file)
wb.active = 0
sheet = wb.active

num = 0
m_arr = []
dataX = [datetime.date(year=2020, month=1, day=1), datetime.date(year=2021, month=1, day=1)]
dataXNew = [datetime.date(year=2020, month=1, day=1), datetime.date(year=2021, month=1, day=1), datetime.date(year=2022, month=1, day=1)]
pred = model.XML_CTYPE_ANY
trn = ['2013-05-26']
r2 = r2_score(trn, pred[1:32])
print('R^2: %1.2f' % r2)
metrics.rmse(trn, pred[1:32])
def plot():
    fig = Figure(figsize=(5, 5), dpi=100)
    y = m_arr
    x = dataX
    plot1 = fig.add_subplot(111)
    plot1.plot(x, y)
    canvas = FigureCanvasTkAgg(fig, master=window)
    canvas.draw()
    canvas.get_tk_widget().pack()
    toolbar = NavigationToolbar2Tk(canvas, window)
    toolbar.update()
    canvas.get_tk_widget().pack()

def printInf():
    comboY.current(1)
    m_arr.append(int(txtSick.get()))
    return m_arr

def restart():
    python = sys.executable
    os.execl(python, python, * sys.argv)

def addChart2022():
    dataX.append(datetime.date(year=2022, month=1, day=1))
    sArr = sum(m_arr) / len(m_arr)
    m_arr.append(sArr)
    return print(dataX, m_arr)

window = Tk()
window.title("Аналитика заболевших Covid-19")
window.geometry('500x600')

btnAdd = Button(master=window, text="Рестарт", command=restart)
btnAdd.pack()

# lblAr = Label(master=window, text="Область", font=("Arial Bold", 10))
# lblAr.pack()
# txtArea = Entry(master=window, width=20)
# txtArea.pack()

lbl2 = Label(master=window, text="Количество заболевших", font=("Arial Bold", 10))
lbl2.pack() #27924 / 42245
txtSick = Entry(master=window, width=20)
txtSick.pack()

lblYears = Label(master=window, text="Год", font=("Arial Bold", 10))
lblYears.pack()
comboY = Combobox(master=window)
comboY['values'] = (2020, 2021)
comboY.current(0)
comboY.pack()

btnAdd = Button(master=window, text="Записать", command=printInf)
btnAdd.pack()

btnAdd = Button(master=window, text="Добавить аналитику на 2022 год", command=addChart2022)
btnAdd.pack()

btnPlot = Button(master=window,
                     command=plot,
                     height=2,
                     width=10,
                     text="График")
btnPlot.pack()
window.mainloop()

